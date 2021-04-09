import bert_score
from bert_score import BERTScorer
import logging
import transformers
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl import Workbook, load_workbook
from os import path
import os
import pandas as pd
import numpy as np

transformers.tokenization_utils.logger.setLevel(logging.ERROR)
transformers.configuration_utils.logger.setLevel(logging.ERROR)
transformers.modeling_utils.logger.setLevel(logging.ERROR)

rcParams["xtick.major.size"] = 0
rcParams["xtick.minor.size"] = 0
rcParams["ytick.major.size"] = 0
rcParams["ytick.minor.size"] = 0

rcParams["axes.labelsize"] = "large"
rcParams["axes.axisbelow"] = True
rcParams["axes.grid"] = True

def score_results(file_to_score):
    
    xlsx = ".xlsx"
    file_path = file_to_score + xlsx

    results_sheet = "Sheet1"
    result_list = list()

    # If the file exists we will open and read from the file
    if path.exists(file_path):

        results_file = load_workbook(file_path)
        
        # If the sheet name that contains the raw results is in the list, we will access that sheet
        if results_sheet in results_file.sheetnames:

            # Get desired worksheet
            raw_results_sheet = results_file[results_sheet]

            row_index = 0
            # Create list containing each row from excel spreadsheet.
            for row in raw_results_sheet.values:
                # print(row)
                # Collect column names from first row
                if row_index == 0:
                    temp_list = list()
                    for value in row:
                        temp_list.append(value)
                    row_index += 1
                    result_list.append(temp_list)
                # For each row, append value to list at appropriate key in results_dict
                elif row_index > 0:
                    temp_list = list()
                    for value in row:
                        temp_list.append(value)
                    result_list.append(temp_list)

        # end if raw_results_sheet in results_file.sheetnames
        
    # end if file exists
    else:
        print("\nno file\n")
        
    print("\nresult_list[0]:\n")
    print(result_list[0])
    
    # Create index of column names
    column_name_index = dict()
    column_index = 0
    for value in result_list[0]:
        
        column_name_index[value] = column_index
        column_index += 1
        
    # end for value in result_list[0]
    
    print("\ncolumn_name_index:\n")
    print(column_name_index)
    
    # Split results into candidate (generated answers) and references (context from which answers were 
    # generated).
    candidates = list()
    references = list()
    
    
    # We skip the first line because it contains column headers
    for i in range (1, len(result_list)):
        
        candidates.append(result_list[i][column_name_index["G_answer"]])
        references.append(result_list[i][column_name_index["Answer"]])
    
    # end for i in range (1, len(result_list))
    
    print("\ncandidate[0]:\n")
    print(candidates[0])
        
    print("\references[0]:\n")
    print(references[0])
    
    BERTScore_dict = dict()
    
    P, R, F1 = scorer.score(candidates, references)
    
    # For each result put an entry in the dictionary with the following.
    for i in range(0, len(result_list) - 1):
    
        # print("\ni:\n{}".format(i))
        BERTScore_dict[i] = {"context": references[i], "generated_answer": candidates[i], "bert_p":P[i], "bert_r":R[i], "bert_f1":F1[i] }
        
    # end for i in range(0, len(result_list))
    
    
    # Convert each tensor value into a float.
    for i in range(0, len(result_list) - 1):
    
        BERTScore_dict[i]["bert_p"] = float(BERTScore_dict[i]["bert_p"])
        BERTScore_dict[i]["bert_r"] = float(BERTScore_dict[i]["bert_r"])
        BERTScore_dict[i]["bert_f1"] = float(BERTScore_dict[i]["bert_f1"])
        
    # end for i in range(0, len(result_list))
    
    print("\nBERTScore_dict[0]:\n")
    print(BERTScore_dict[0])
    
    
    # Write all results to .xlsx file
    BERTScored_file = file_to_score + "_BERTScore" + ".xlsx"
    print(BERTScored_file)
    
    final_data = np.array([['context', 'generated_answer', 'bert_p', 'bert_r', 'bert_f1']])
    
    for entry in BERTScore_dict.keys():
        
        temp_data = [BERTScore_dict[entry]["context"], BERTScore_dict[entry]["generated_answer"], BERTScore_dict[entry]["bert_p"], BERTScore_dict[entry]["bert_r"], BERTScore_dict[entry]["bert_f1"]]
        final_data = np.append(final_data, [temp_data], axis=0)
    
    df = pd.DataFrame(final_data)
    df.to_excel(BERTScored_file, index = False, header= False)
    
# end def score_results

print("Preparations complete...")


scorer = BERTScorer(lang="en", rescale_with_baseline=True)


files_to_score = [
    
    "/home/bdlabucdenver/data/five_answer_best_context_concat",
    "/home/bdlabucdenver/data/four_best_context_concat",
    "/home/bdlabucdenver/data/two_answer_best_context_concat",
    "/home/bdlabucdenver/data/two_answer_two_context_concat"
    
]

for file in files_to_score:
    score_results(file)

print("We are done here...")
